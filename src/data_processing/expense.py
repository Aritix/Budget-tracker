from src.data_processing.entry import Entries, Entry
from src.data_processing.rule import Rules, Rule
import datetime
import matplotlib.pyplot as plt
import numpy as np
from src.data_processing.constantes import YEAR, CURRENCY
from random import shuffle
from openpyxl import Workbook
import json
from typing import Self
import yaml
from yaml import CLoader as Loader, CDumper as Dumper
import xlsxwriter


class Expenses:
    """
    Object to track expenses and incomes based on the *Entry* and *Entries* objects.
    """

    default_class = "Autres"
    colors = [
        "#005f73",
        "#0081a7",
        "#00b4d8",
        "#90e0ef",
        "#48cae4",
        "#b5838d",
        "#a3b18a",
        "#6d6875",
        "#555b4c",
        "#2b2d42",
        "#370617",
        "#6a040f",
    ]

    def __init__(self, default_class="Autres"):
        self.rules: Rules = Rules()
        self.entries: Entries = Entries()
        self.default_class = default_class
        self.sums = {default_class: 0}
        self.categories = []

    def add_expense(self, entry: Entry, category_name: str):
        """
        Update the object with the new categorized entry
        """
        entry.category = category_name
        if category_name not in self.categories:
            self.categories.append(category_name)
        entry.price = round(entry.price, 2)
        # self.entries.add_entry_object(entry)
        self.sums[category_name] = entry.price + (
            self.sums[category_name] if category_name in self.sums.keys() else 0
        )

    def show_by_category(self):
        str_res = ""
        for category, amount in self.sums.items():
            str_res += f"{category}, {amount}\n"
        return str_res

    def to_spreadsheet(self):
        """
        Export all expenses to a workbook in "expenses.xlsx"
        """
        # Workbook creation
        workbook = xlsxwriter.Workbook("expenses.xlsx")
        data_sheet = workbook.add_worksheet("RawData")
        insight_sheet = workbook.add_worksheet("Insights")
        date_format = workbook.add_format({"num_format": "yyyy-mm-dd"})

        # Writing raw data
        data_sheet.write(0, 0, "Date")
        data_sheet.write(0, 1, "Month")
        data_sheet.write(0, 2, "Amount")
        data_sheet.write(0, 3, "Description")
        data_sheet.write(0, 4, "Category")
        data = []
        for i, entry in enumerate(self.entries):
            data.append(
                [
                    f"2025-{entry.month}-{entry.day}",
                    f"2025-{entry.month}",
                    entry.price,
                    entry.description,
                    entry.category,
                ]
            )
        options = {
            "data": data,
            "columns": [
                {"header": "Date", "total_string": "Totals", "format": date_format},
                {"header": "Month", "total_string": "Totals"},
                {
                    "header": "Amount",
                    "total_function": f"sum(B1:B{len(self.entries)+1})",
                },
                {"header": "Description"},
                {"header": "Category"},
            ],
        }
        data_sheet.add_table(f"A1:E{len(self.entries)+2}", options)

        workbook.close()

    def export_to_spreadsheet(self, filename="output_formulas.xlsx"):
        """
        Export the *Expenses* object to a .xlsx spreadsheet format.
        """
        wb = Workbook()
        sheet = wb.active

        # Write data and labels
        sheet["A1"] = "Date"
        sheet["B1"] = "Amount"
        sheet["C1"] = "Description"
        sheet["D1"] = "Category"
        for i, entry in enumerate(self.entries, start=2):
            sheet.cell(row=i, column=1, value=f"{entry.day}/{entry.month}")
            sheet.cell(row=i, column=2, value=entry.price)
            sheet.cell(row=i, column=3, value=entry.description)
            sheet.cell(row=i, column=4, value=entry.category)

        # Write formulas (e.g., SUMIF)
        sheet["E1"] = "Sum of A"
        sheet["E2"] = '=SUMIF(A2:A{}, "A", B2:B{})'.format(
            len(self.entries) + 1, len(self.entries) + 1
        )
        sheet["F1"] = "Sum of B"
        sheet["F2"] = '=SUMIF(A2:A{}, "B", B2:B{})'.format(
            len(self.entries) + 1, len(self.entries) + 1
        )

        wb.save(filename)

    def time_graph(self, subplot=None, SavedFileName=None, show=False):
        labels = self.time_sums.keys()
        items = sum([self.time_sums[label] for label in labels], [])
        times = [tpl[0] for tpl in items]
        times = [
            datetime.datetime.strptime(f"{YEAR}-{tpl[1]}-{tpl[0]}", "%Y-%m-%d")
            for tpl in times
        ]
        bottom_dict = {}
        for time in times:
            bottom_dict[time] = 0
        ax = plt.subplot(111) if subplot == None else subplot
        for (category_name, list_expenses), color in zip(
            self.time_sums.items(), self.colors
        ):
            times = [
                datetime.datetime.strptime(
                    f"{YEAR}-{tpl[0][1]}-{tpl[0][0]}", "%Y-%m-%d"
                )
                for tpl in list_expenses
            ]
            amounts = [tpl[1] for tpl in list_expenses]
            # Listes aggrégées pour fusionner les dépenses de même catégorie à la même date.
            times_agg = []
            amounts_agg = []
            for time, amount in zip(times, amounts):
                if time in times_agg:
                    amounts_agg[times_agg.index(time)] += amount
                else:
                    times_agg.append(time)
                    amounts_agg.append(amount)
            bottom = [bottom_dict[time] for time in times_agg]
            ax.bar(
                times_agg,
                amounts_agg,
                label=category_name,
                bottom=bottom,
                color=color,
            )
            for time, amount in zip(times_agg, amounts_agg):
                bottom_dict[time] += amount

        ax.grid()
        ax.set_title(
            f"Dépense du {self.time_min.strftime("%m/%d")} au {self.time_max.strftime("%m/%d")}"
        )
        ax.legend()
        ax.set_xlabel("Date")
        ax.set_ylabel(f"Montant ({CURRENCY})")
        if SavedFileName:
            plt.savefig(SavedFileName)
        else:
            if show:
                plt.show()

    def income_pie(self, subplot=None, SavedFileName=None, show=False):
        labels = self.sums.keys()
        amounts = [(-self.sums[name] if self.sums[name] <= 0 else 0) for name in labels]
        ax = plt.subplot(111) if subplot == None else subplot

        def make_autopct(values):
            def my_autopct(pct):
                total = sum(values)
                val = int(round(pct * total / 100.0))
                # return "{p:.2f}%  ({v:d}{c})".format(p=pct, v=val, c=" " + CURRENCY)
                return f"{val:d}"

            return my_autopct

        # shuffle(colors)
        wedges, texts, autotests = ax.pie(
            amounts,
            labels=labels,
            autopct=make_autopct(amounts),
            pctdistance=0.4,
            explode=[1 / (len(amounts) * 10) for _ in amounts],
            colors=self.colors,
            shadow=False,
            startangle=0,
            radius=1,
            wedgeprops=dict(width=0.5, mouseover=True),
            # textprops=dict(color='white')
        )
        for i, (wedge, text, autotext) in enumerate(zip(wedges, texts, autotests)):
            text.set_color(wedge.get_facecolor())
            text.set_fontweight("bold")
            text.set_fontsize("large")
            text.set_multialignment("left")
            autotext.set_color(wedge.get_facecolor())
            autotext.set_fontweight("bold")

        plt.title("Répartition des dépenses")
        if SavedFileName:
            plt.savefig(SavedFileName)
        else:
            if show:
                plt.show()

    def pie_graph(self, subplot=None, SavedFileName=None, show=False):
        labels = self.sums.keys()
        amounts = [(self.sums[name] if self.sums[name] >= 0 else 0) for name in labels]
        ax = plt.subplot(111) if subplot == None else subplot

        def make_autopct(values):
            def my_autopct(pct):
                total = sum(values)
                val = int(round(pct * total / 100.0))
                # return "{p:.2f}%  ({v:d}{c})".format(p=pct, v=val, c=" " + CURRENCY)
                return f"{val:d}"

            return my_autopct

        # shuffle(colors)
        wedges, texts, autotests = ax.pie(
            amounts,
            labels=labels,
            autopct=make_autopct(amounts),
            pctdistance=0.4,
            explode=[1 / (len(amounts) * 10) for _ in amounts],
            colors=self.colors,
            shadow=False,
            startangle=0,
            radius=1,
            wedgeprops=dict(width=0.5, mouseover=True),
            # textprops=dict(color='white')
        )
        for i, (wedge, text, autotext) in enumerate(zip(wedges, texts, autotests)):
            text.set_color(wedge.get_facecolor())
            text.set_fontweight("bold")
            text.set_fontsize("large")
            text.set_multialignment("left")
            autotext.set_color(wedge.get_facecolor())
            autotext.set_fontweight("bold")

        plt.title("Répartition des dépenses")
        if SavedFileName:
            plt.savefig(SavedFileName)
        else:
            if show:
                plt.show()

    def all_graph(self, SavedFileName=None, show=False):
        fig = plt.figure(figsize=(15, 6))
        self.income_pie(plt.subplot(131), show=False)
        self.pie_graph(plt.subplot(132), show=False)
        self.time_graph(plt.subplot(133), show=False)
        fig.autofmt_xdate()
        if SavedFileName:
            plt.savefig(SavedFileName)
        else:
            if show:
                plt.show()

    def oldSave(self, filename="outputs.json", tostring=False) -> str | None:
        """
        Save the *Expenses* objects as a JSON string. Can be later loaded by the *load* class method.
        """
        obj = {
            "default_class": self.default_class,
            "entries": self.entries.save(),
        }
        if filename:
            with open(filename, "w") as file:
                file.write(json.dumps(obj))
        if tostring:
            return json.dumps(obj)

    def save(self, filename="expenses.yml") -> None:
        res = yaml.dump(self, Dumper=Dumper)
        if filename:
            with open(filename, "w") as f:
                f.write(res)
        return res

    def get_category_json(self) -> dict:
        """
        Get the categories and sums by category of the *Expenses* object.
        Only the positive entries for now.
        """
        positive_sums = {
            category: amount for category, amount in self.sums.items() if amount > 0
        }
        negative_sums = {
            category: amount for category, amount in self.sums.items() if amount < 0
        }
        return positive_sums, negative_sums

    def get_months_json(self) -> dict:
        """
        Get the months and sums by month of the *Expenses* object.
        Only the positive entries for now.
        """
        all_months = []
        # months_sums = {}
        positive_sums = {}
        negative_sums = {}
        entry: Entry = None
        for entry in self.entries:
            if entry.month not in all_months:
                all_months.append(entry.month)
                positive_sums[entry.month] = 0
                negative_sums[entry.month] = 0
            if entry.price > 0:
                positive_sums[entry.month] += entry.price
            else:
                negative_sums[entry.month] += entry.price
        return positive_sums, negative_sums
        ## New format
        income_json = {}
        outcome_json = {}
        months_json = {"incomes": income_json, "outcomes": outcome_json}
        month = ""
        transaction_direction = ""
        for entry in self.entries:
            month = entry.month
            if entry.price > 0:
                json_to_fill = outcome_json
            elif entry.price < 0:
                json_to_fill = income_json
            if month not in json_to_fill:
                json_to_fill[month] = 0
            json_to_fill[month] += entry.price
        print(months_json)

        return months_sums

    def filter(self, time_filter: list[str], category_filter: list[str]) -> Self:
        """
        Filter the entries of the *Expenses* object based on given filters (by category name and time).
        """
        new_expenses = Expenses()
        new_expenses.set_rules(self.rules)
        for entry in self.entries:
            # Category filter
            if (
                entry.category in category_filter or "*" in category_filter
            ):  # and entry.day in time_filter:
                if (
                    entry.month in time_filter or "*" in time_filter
                ):  # and entry.day in time_filter:
                    new_expenses.add_expense(entry, entry.category)
                    new_expenses.entries.add_entry_object(entry)
        return new_expenses

    # Rules methods

    def set_rules(self, rules: Rules) -> None:
        """
        Set *Rules* object and update categories.
        """
        self.rules = rules
        self.categories = []
        for rule in rules:
            if rule.classe_name not in self.categories:
                self.categories.append(rule.classe_name)

    def add_rule_fields(
        self, new_category_name: str, index: int, patterns: list[str]
    ) -> None:
        """
        Add a new rule to the Rules object and update categories.
        """
        self.rules.add_rule_fields(new_category_name, index, patterns)
        if new_category_name not in self.categories:
            self.categories.append(new_category_name)

    def delete_rule(self, category_name: str) -> None:
        """
        Delete a rule from the Rules objects and updates categories.
        """
        self.rules.delete_rule(category_name)
        if category_name in self.categories:
            self.categories.remove(category_name)

    def edit_rule(self, category_name: str, changes: dict) -> None:
        """
        Edit the rules of the category specified. For each key in *changes*, changed this key as as pattern (if present) to its value.
        """
        self.rules.edit_rule_pattern(category_name, changes)


    def getUncategorizedEntries(self) -> Entries:
        filtered_entries = Entries()
        for entry in self.entries:
            if entry.category == self.default_class or entry.category is None:
                filtered_entries.add_entry_object(entry)
        return filtered_entries

    def updateCategorization(self, only_uncategorized: bool = True) -> None:
        expenses_parser = Expenses_parser()
        expenses_parser.load(self.entries, self.rules, self)
        expenses_parser.parse(update_categorized=not only_uncategorized)

    def __str__(self):
        return f"<Expenses object with total amount of {sum(self.sums.values())}>"

    def __getstate__(self):
        # Return a dictionary containing the attributes you want to serialize
        return {
            "rules": self.rules,
            "entries": self.entries,
            "default_class": self.default_class,
            "sums": self.sums,
            "categories": self.categories,
        }

    def __setstate__(self, state):
        # Reconstruct the object from the state dictionary
        self.rules = state["rules"]
        self.entries = state["entries"]
        self.default_class = state["default_class"]
        self.sums = state["sums"]
        self.categories = state["categories"]

    @classmethod
    def oldLoad(cls, filename="outputs.json") -> Self:
        """
        Load an *Expenses* object saved with the *save* method.
        """
        with open(filename, "r") as file:
            obj = json.load(file)
        expenses = Expenses(default_class=obj["default_class"])
        entries = Entries.load(obj["entries"].strip())
        for entry in entries:
            expenses.add_expense(entry, entry.category)
        return expenses

    @classmethod
    def load(cls, filename="expenses.yml") -> Self:
        # yaml.add_representer(Expenses, expenses_representer)
        # yaml.add_constructor(u'!Expenses', expenses_constructor)
        with open(filename, "r") as f:
            stream = f.read()
        return yaml.load(stream, Loader=Loader)
